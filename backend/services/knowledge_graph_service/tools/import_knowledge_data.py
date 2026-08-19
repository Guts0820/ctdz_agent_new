"""Idempotently import the existing knowledge and error-bank assets into Neo4j."""

import csv
import sqlite3
from pathlib import Path
import sys
from typing import Any

SERVICE_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = SERVICE_DIR.parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

from backend.services.knowledge_graph_service.database import neo4j_conn


KNOWLEDGE_POINTS_PATH = PROJECT_ROOT / "database" / "seed" / "knowledge_points.csv"
KNOWLEDGE_EXPLANATIONS_PATH = PROJECT_ROOT / "database" / "seed" / "knowledge_explanations.csv"
ERROR_REFERENCE_PATH = PROJECT_ROOT / "database" / "reference" / "三级错因标签.xlsx"
SQLITE_PATH = PROJECT_ROOT / "database" / "sqlite" / "example_db.db"


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _grade(value: object) -> int | None:
    text = str(value or "").strip()
    try:
        return int(text) if text else None
    except ValueError:
        return None


def load_knowledge_points(
    points_path: Path = KNOWLEDGE_POINTS_PATH,
    explanations_path: Path = KNOWLEDGE_EXPLANATIONS_PATH,
    sqlite_path: Path = SQLITE_PATH,
) -> list[dict[str, Any]]:
    explanations = {
        row["knowledge_id"]: row
        for row in _read_csv(explanations_path)
        if row.get("knowledge_id")
    }
    sqlite_metadata: dict[str, sqlite3.Row] = {}
    try:
        with sqlite3.connect(sqlite_path) as connection:
            connection.row_factory = sqlite3.Row
            sqlite_metadata = {
                row["knowledge_id"]: row
                for row in connection.execute("SELECT * FROM knowledge").fetchall()
            }
    except sqlite3.Error:
        sqlite_metadata = {}
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in _read_csv(points_path):
        knowledge_id = str(row.get("knowledge_point_id", "")).strip()
        if not knowledge_id or knowledge_id in seen:
            raise ValueError(f"知识点 ID 为空或重复：{knowledge_id!r}")
        seen.add(knowledge_id)
        detail = explanations.get(knowledge_id, {})
        metadata = sqlite_metadata.get(knowledge_id)
        title = (detail.get("title") or row.get("description") or "").strip()
        items.append(
            {
                "id": knowledge_id,
                "title": title,
                "description": row.get("description", "").strip(),
                "grade": _grade(row.get("grade")),
                "semester": row.get("semester", "").strip(),
                "content": detail.get("content", "").strip(),
                "key_formulas": detail.get("key_formulas", "").strip(),
                "common_mistakes": detail.get("common_mistakes", "").strip(),
                "teaching_points": detail.get("teaching_points", "").strip(),
                "difficulty": metadata["difficulty"] if metadata else "",
                "textbook_version": metadata["textbook_version"] if metadata else row.get("version", ""),
                "unit": metadata["unit"] if metadata else "",
                "prerequisite": metadata["prerequisite"] if metadata else "",
                "next_knowledge": metadata["next_knowledge"] if metadata else "",
                "is_core": bool(metadata["is_core"]) if metadata else False,
                "source": "database/seed",
            }
        )
    if not items:
        raise ValueError("知识点数据为空")
    return items


def _load_errors_from_sqlite(path: Path = SQLITE_PATH) -> list[dict[str, Any]]:
    with sqlite3.connect(path) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            "SELECT error_id, level1, level2, level3, error_description FROM error_bank"
        ).fetchall()
    return [
        {
            "id": row["error_id"],
            "level1": row["level1"] or "",
            "level2": row["level2"] or "",
            "level3": row["level3"] or "",
            "criteria": row["error_description"] or "",
            "grade_range": "",
            "knowledge_scope": "",
            "example": "",
            "name": row["level3"] or "",
            "source": "database/sqlite/example_db.db",
        }
        for row in rows
    ]


def load_error_causes(reference_path: Path = ERROR_REFERENCE_PATH) -> list[dict[str, Any]]:
    """Load the richer reviewed Excel asset, falling back to SQLite's error bank."""
    try:
        from openpyxl import load_workbook

        worksheet = load_workbook(reference_path, read_only=True, data_only=True).active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("错因参考表为空")
        headers = [str(value or "").strip() for value in rows[0]]
        index = {header: position for position, header in enumerate(headers)}
        required = {"错因ID", "一级标签", "二级标签", "三级细分错因"}
        if not required.issubset(index):
            raise ValueError("错因参考表缺少必要列")
        items = []
        seen: set[str] = set()
        for values in rows[1:]:
            get = lambda name: str(values[index[name]] or "").strip() if index[name] < len(values) else ""
            item = {
                "id": get("错因ID"),
                "level1": get("一级标签"),
                "level2": get("二级标签"),
                "level3": get("三级细分错因"),
                "criteria": get("判定标准"),
                "grade_range": get("适用学段"),
                "knowledge_scope": get("知识点范围"),
                "example": get("典型示例"),
                "name": get("三级细分错因"),
                "source": "database/reference/三级错因标签.xlsx",
            }
            if not item["id"] or item["id"] in seen:
                raise ValueError(f"错因 ID 为空或重复：{item['id']!r}")
            seen.add(item["id"])
            items.append(item)
        if items:
            return items
    except (ImportError, OSError, ValueError):
        pass
    return _load_errors_from_sqlite()


def import_knowledge_points(items: list[dict[str, Any]]) -> int:
    neo4j_conn.query(
        "CREATE CONSTRAINT knowledge_point_id_unique IF NOT EXISTS "
        "FOR (k:KnowledgePoint) REQUIRE k.id IS UNIQUE"
    )
    result = neo4j_conn.query(
        """
        UNWIND $items AS item
        MERGE (k:KnowledgePoint {id: item.id})
        SET k.title = item.title,
            k.description = item.description,
            k.grade = item.grade,
            k.semester = item.semester,
            k.content = item.content,
            k.key_formulas = item.key_formulas,
            k.common_mistakes = item.common_mistakes,
            k.teaching_points = item.teaching_points,
            k.difficulty = item.difficulty,
            k.textbook_version = item.textbook_version,
            k.unit = item.unit,
            k.prerequisite = item.prerequisite,
            k.next_knowledge = item.next_knowledge,
            k.is_core = item.is_core,
            k.source = item.source,
            k.import_batch = 'knowledge-assets-v1'
        RETURN count(k) AS imported_count
        """,
        {"items": items},
    )
    return int(result[0]["imported_count"]) if result else 0


def import_error_causes(items: list[dict[str, Any]]) -> int:
    neo4j_conn.query(
        "CREATE CONSTRAINT error_cause_id_unique IF NOT EXISTS "
        "FOR (e:ErrorCause) REQUIRE e.id IS UNIQUE"
    )
    result = neo4j_conn.query(
        """
        UNWIND $items AS item
        MERGE (e:ErrorCause {id: item.id})
        SET e.level1 = item.level1,
            e.level2 = item.level2,
            e.level3 = item.level3,
            e.criteria = item.criteria,
            e.grade_range = item.grade_range,
            e.knowledge_scope = item.knowledge_scope,
            e.example = item.example,
            e.name = item.name,
            e.source = item.source,
            e.import_batch = 'knowledge-assets-v1'
        RETURN count(e) AS imported_count
        """,
        {"items": items},
    )
    neo4j_conn.query(
        """
        MATCH (e:ErrorCause), (k:KnowledgePoint)
        WHERE e.knowledge_scope <> ''
          AND (e.knowledge_scope = k.title
            OR k.title CONTAINS e.knowledge_scope
            OR e.knowledge_scope CONTAINS k.title)
        MERGE (e)-[:APPLIES_TO]->(k)
        """
    )
    return int(result[0]["imported_count"]) if result else 0


def main() -> None:
    knowledge = load_knowledge_points()
    errors = load_error_causes()
    knowledge_count = import_knowledge_points(knowledge)
    error_count = import_error_causes(errors)
    print(f"已导入或更新知识点 {knowledge_count} 条、错因 {error_count} 条。")


if __name__ == "__main__":
    main()
